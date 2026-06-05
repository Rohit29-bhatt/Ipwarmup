"""
Run this once to generate a sample_credentials.xlsx for testing.
Usage: python generate_sample.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Credentials"

# Header style
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1e2535", end_color="1e2535", fill_type="solid")

headers = ["email", "password", "subject_line"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 45

# Sample rows
rows = [
    ["testuser1@gmail.com",   "password123",  "Q3 Newsletter — Important Update"],
    ["testuser2@outlook.com", "securepass456", "Q3 Newsletter — Important Update"],
    ["testuser3@yahoo.com",   "mypassword789", ""],   # will use global subject
    ["testuser4@hotmail.com", "hotpass321",    "Special Offer Inside!"],
    ["testuser5@gmail.com",   "gmailpass654",  "Q3 Newsletter — Important Update"],
]
for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)

wb.save("sample_credentials.xlsx")
print("✓ sample_credentials.xlsx created successfully!")
print("  Upload this file to the IP Warmup tool to get started.")
